import json
from json2html import *

with open("test.json") as f:  # give your file input path
    d = json.load(f)  # load our input file to d
    scanOutput = json2html.convert(json=d)
    print(  scanOutput)
    htmlReportFile = "result.html"
    with open (htmlReportFile, 'w') as htmlfile:
        htmlfile.write(str(scanOutput))
        
        
        ws1.column_dimensions['A'].width=10
ws1.row_dimensions[1].height=20

ws1['A1']='Welcome '
ws1['B1']='to plus2net'
d=ws1.cell(2,1,'Python')

def home(request):
    return HttpResponse(" THIS IS A HOME PAGE OF THIS APP ")
    
    def home(request):
    return HttpResponse(" THIS IS A HOME PAGE OF THIS APP ")

print("key: " + str(x))
    print("value: " + str(y))
    
    
ws1['A2']='PYTHON'
ws1['B2']=3.0
ws1['A3']='JAVA'
ws1['B3']=9.


for x,y in d.items():
    print(x)
    for i in (y):
        print (i)


MEDIA_URL='media/'

MEDIAFILES_DIRS = BASE_DIR, 'MEDIA'


        for j in l:
            c=j['value']
            if c == 'Server-Linux':
                print(c)
                
                
                
                
            if 'linux' in tmp_name or 'windows' in tmp_name:
                ws1['B'+str(counter)] = (j['name'])
                break
            #else:
                #ws1['B' + str(counter)] = (j['name'])

                m=+1

            for k in i['tags']:
                if k['value'] == 'Server-Linux' or 'Server-Windows':
                    ws1['C' + str(counter)] = (k['value'])
                    
                    
                    
                            for j in i['parentTemplates']:
            tmp_name = j['name']
            tmp_name=tmp_name.lower()
            #if j['name'] == 'Template OS Linux by Zabbix agent' or 'Template OS Windows by Zabbix agent':
            print(tmp_name)
            
---------------------------------------------------------------------------------------------------------------------------------------------------

# ALL SERVERS ADDED IN ZABBIX WITH TAGS AND TEMPLATE 

import json
from openpyxl import Workbook
wb = Workbook()
ws1 = wb.active
with open("templates.json") as f:
    d = json.load(f)
    z=d['result']
    #HC = 1
    counter = 2
    temp_row = 2
    tags_row = 2
    for i in z:
        # l=i['host']
        ws1['A' + str(counter)] = (i['host'])
        m = i['parentTemplates']

        temp_col = 2
        for j in m:
            x4 = ws1.cell(row=temp_row, column=temp_col)
            x4.value = j['name']
            temp_col= temp_col + 1

            tags_col = 5
            for k in i['tags']:
                print(k['value'])
                x5 = ws1.cell(row=tags_row, column=tags_col)
                x5.value = k['value']
                tags_col = tags_col + 1

        temp_row = temp_row + 1
        tags_row = tags_row +1
        counter = counter + 1

wb.save("templates.xlsx")

----------------------------------------------------------------------------------------------------------------------------------------------------

{% block content %}


{% endblock 

{% include 'slider.html' with sliderdata=slider %}
slide{
    width: 50%;
    height: 50%;
    position: absolute;
}

=\-------------
        c2=1
        c3=1

        m = i['tags']
        for j in m:
            if j['tag'] == 'CRMID':
                x2 = ws1.cell(row=r2, column=c2)
                x2.value = (j['value'])

            elif j['tag'] == 'CustomerName':
                x3 = ws1.cell(row=r3, column=c3)
                x3.value = (j['value'])
        c2=c2+1
        c3=c3+1
        
        
        engine = create_engine("mysql://root:Sarvesh@21@localhost:3306/mydb")
        
       
from sqlalchemy import create_engine
from sqlalchemy import text

engine = create_engine("mysql://root:SArvesh@21@localhost3306/mydb", echo=True)
conn = engine.connect()
sql_text = text("SELECT * FROM mydb.employee")
result = conn.execute(sql_text)

for row in result:
    print(row)
    
-----------------------------------------------------------------------------------------------------------------------------------   
  
import pymysql

dbServerName = "127.0.0.1"
dbUser = "root"
dbPassword = "Sarvesh@21"
dbName = "mydb"
charSet = "utf8mb4"
cusrorType = pymysql.cursors.DictCursor

connectionObject = pymysql.connect(host=dbServerName, user=dbUser, password=dbPassword, db=dbName, charset=charSet,
                                   cursorclass=cusrorType)
# Create a cursor object
cursorObject = connectionObject.cursor()
# SQL query string
sqlQuery = "select * from employee"
# Execute the sqlQuery
cursorObject.execute(sqlQuery)
# Fetch all the rows
rows = cursorObject.fetchall()
for row in rows:
    print(row)



from flask import Flask
from flask_sqlalchemy import SQLAlchemy

app = Flask(__name__)

app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql://root:@localhost/mydb'
db = SQLAlchemy(app)

class Employee(db.Model):
    EmpID = db.Column(db.Integer, primary_Key=True)
    EmpName = db.Column(db.String(80), unique=False, nullable=False)
    EmpAge = db.Column(db.Integer, primary_Key=False, nullable=False)
    Dept = db.Column(db.String(80), unique=False, nullable=False)


if __name__ == 'main':
    app.run(debug=True)
print(engine)



import pymysql

connection = pymysql.connect(host='localhost',
                             user='root',
                             password='Sarvesh@21',
                             db='mydb')

cursor = connection.cursor()

# Create a new record
sql = "INSERT INTO employee (EmpID, EmpName, EmpAge, Dept) VALUES (%s, %s, %s, %s)"

# Execute the query
cursor.execute(sql, (3, 'Kabir', 25, 'CLOUD'))

# the connection is not autocommited by default. So we must commit to save our changes.
connection.commit()